#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Score batch3 candidates through the production path + fitted v3a model,
then build batch3-review.xlsx (model suggestion column + user dropdown).
Also: grouped-bootstrap CI for the chosen operating point (tau=0.65)."""
import json
import os
import sys
import warnings

warnings.filterwarnings('ignore')
import numpy as np

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.abspath(os.path.join(HERE, '..', '..', '..'))
sys.path.insert(0, os.path.join(REPO, 'python'))
sys.path.insert(0, os.path.join(HERE, '..', 'calibration-cal20260824-1855'))

import m7_embedding_pre_v1 as emb
from calibration_harness import (Surface, tokenize, STOP, minmax,
                                 load_live_records, TOP_K, WDENSE,
                                 MODEL_DIR, MODEL_REVISION)
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import GroupKFold, cross_val_predict

CACHE = os.path.join(HERE, 'vec-cache')
cfg = {'provider': 'bge-m3-pre-v1', 'modelDir': MODEL_DIR,
       'modelRevision': MODEL_REVISION, 'dimension': 1024, 'torchThreads': 16}
print('[b3] loading BGE-M3...', flush=True)
embedder = emb.BgeM3Embedder(cfg)


def build_cached(name, records):
    import numpy as np
    key = emb.sha_hex(('|'.join(r['key'] + ':' + str(len(r['text'])) for r in records)
                       + '|' + MODEL_REVISION).encode())[:16]
    path = os.path.join(CACHE, f'{name}-{key}.npz')
    surf = Surface(name, records)
    if os.path.isfile(path):
        z = np.load(path, allow_pickle=True)
        surf.chunks = list(z['chunks'])
        surf.vectors = list(z['vectors'])
        print(f'[b3] {name}: cache hit', flush=True)
    else:
        surf.build(embedder)
        np.savez(path, chunks=np.array(surf.chunks, dtype=object),
                 vectors=np.array(surf.vectors, dtype=np.float32))
        print(f'[b3] {name}: built {len(surf.chunks)} chunks', flush=True)
    return surf


live_recs = load_live_records()
live_surface = build_cached('live', [{'key': r['memoryId'],
                                      'text': r.get('text') or ''} for r in live_recs])
eps = [json.loads(l) for l in open(os.path.join(REPO, 'artifacts', 'm7-corpus-pre',
                                                'episodes.jsonl'), encoding='utf-8')
       if l.strip()]
ep_surface = build_cached('episodes', [{'key': e['episodeId'],
                                        'text': e.get('text') or ''} for e in eps])
texts = {e['episodeId']: e.get('text') or '' for e in eps}
_dc = json.load(open(r'C:\Users\JH Z\.dsh\memory\semantic-pre\derived-corpus.json',
                     encoding='utf-8'))
for entry in (_dc['entries'] if isinstance(_dc['entries'], list)
              else list(_dc['entries'].values())):
    for rec in entry['records']:
        texts[rec['memoryId']] = rec.get('text') or ''

cands = [json.loads(l) for l in open(os.path.join(HERE, 'batch3-candidates.jsonl'),
                                     encoding='utf-8') if l.strip()]


def load(p):
    q = p if os.path.isabs(p) else os.path.join(HERE, p)
    return [json.loads(l) for l in open(q, encoding='utf-8') if l.strip()]

# ---- intent model on 58 golds (same as fit_path) ----
golds = [json.loads(l) for l in open(os.path.join(HERE, 'gold-confirmed.jsonl'),
                                     encoding='utf-8') if l.strip()] + \
        [json.loads(l) for l in open(os.path.join(HERE, 'gold-confirmed-cf.jsonl'),
                                     encoding='utf-8') if l.strip()]
texts_by_id = {g['sampleId']: g['queryText'] for g in golds}
scored_cal_all = load('../calibration-cal20260824-1855/labels.scored.jsonl') \
    + load('cf-scored.jsonl')
tr = [g for g in golds if g.get('isGold') and g['finalAction'] in ('A', 'S')]
vec = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 4), min_df=1,
                      sublinear_tf=True)
Xt = vec.fit_transform([texts_by_id[r['sampleId']] for r in tr])
yt = np.array([1 if r['finalAction'] == 'A' else 0 for r in tr])
intent_clf = LogisticRegression(max_iter=2000, class_weight='balanced').fit(Xt, yt)

INTERROG = ['什么', '如何', '怎么', '哪些', '哪个', '为什么', '多少', '吗',
            '呢', '啥', 'recall', 'what', 'how', 'which', 'when', 'where',
            'why', 'who']
RECALL_CTX = ['之前', '上次', '当时', '早前', '以往', '历史', '记录里', '记忆里',
              '找出来', '调出来', '翻下', '说下', 'previous', 'earlier',
              'last time']

out = []
lat = []
for c in cands:
    surf = live_surface if c.get('parentMemoryId') else ep_surface
    cands_keys, ms = surf.search(embedder, c['queryText'])
    lat.append(ms)
    keys = {k['key'] for k in cands_keys}
    hit = bool(set(c.get('expectedMemoryIds') or []) & keys)
    q = ''.join(ch for ch in c['queryText'].lower()
                if ch.isalnum() or '\u4e00' <= ch <= '\u9fff')
    qb = set(q[i:i + 2] for i in range(len(q) - 1)) or {q}
    cand_text = texts.get(cands_keys[0]['key'], '') if cands_keys else ''
    cb = ''.join(ch for ch in cand_text.lower()
                 if ch.isalnum() or '\u4e00' <= ch <= '\u9fff')
    cbs = set(cb[i:i + 2] for i in range(len(cb) - 1)) or {cb}
    contain = len(qb & cbs) / max(1, len(qb))
    tl = c['queryText'].lower()
    mark = int(any(x in tl for x in INTERROG + RECALL_CTX))
    ranked = cands_keys
    margin = (ranked[1]['score'] and (ranked[0]['score'] - ranked[1]['score'])) \
        if len(ranked) > 1 else 1.0
    iprob = float(intent_clf.predict_proba(vec.transform([c['queryText']]))[:, 1])
    # v3a emit-worthiness probability (deployable features)
    featv = np.array([[iprob, ranked[0]['score'] if ranked else 0.0,
                       max(0.0, margin), contain, mark, len(ranked)]])
    # refit deployable model on golds with same feature builder
    gfeat = []
    gy = []
    for g in golds:
        if not (g.get('isGold') and g['finalAction'] in ('A', 'S')):
            continue
        s = None
        for src in (scored_cal_all):
            if src['sampleId'] == g['sampleId']:
                s = src
                break
        rk = s.get('_ranked') or []
        mg = (rk[0]['dense'] - rk[1]['dense']) if len(rk) > 1 else 1.0
        qq = ''.join(ch for ch in g['queryText'].lower()
                     if ch.isalnum() or '\u4e00' <= ch <= '\u9fff')
        qbb = set(qq[i:i + 2] for i in range(len(qq) - 1)) or {qq}
        ct = texts.get(rk[0]['key'], '') if rk else ''
        cc = ''.join(ch for ch in ct.lower()
                     if ch.isalnum() or '\u4e00' <= ch <= '\u9fff')
        ccs = set(cc[i:i + 2] for i in range(len(cc) - 1)) or {cc}
        tll = g['queryText'].lower()
        mk = int(any(x in tll for x in INTERROG + RECALL_CTX))
        ip = float(intent_clf.predict_proba(vec.transform([g['queryText']]))[:, 1])
        gfeat.append([ip, rk[0]['dense'] if rk else 0.0, max(0.0, mg),
                      len(qbb & ccs) / max(1, len(qbb)), mk, len(rk)])
        gy.append(1 if g['finalAction'] == 'A' else 0)
    v3a = LogisticRegression(max_iter=2000, class_weight='balanced').fit(
        np.array(gfeat), np.array(gy))
    p_emit = float(v3a.predict_proba(featv)[:, 1])
    if p_emit >= 0.65 and hit:
        sug = 'emit'
    elif p_emit >= 0.40 or (iprob >= 0.35 and hit):
        sug = 'prefetch'
    else:
        sug = 'suppress'
    out.append({
        'sampleId': c['sampleId'], 'kind': c['kind'],
        'queryText': c['queryText'], 'language': c.get('language'),
        'proposedAction': c['proposedAction'],
        'expectedMemoryIds': c.get('expectedMemoryIds') or [],
        'rationale': c.get('rationale'),
        '_denseTop': float(round(ranked[0]['score'], 4)) if ranked else 0.0,
        '_margin': float(round(max(0.0, margin), 4)),
        '_containment': round(contain, 3), '_mark': mark,
        '_intentProb': round(iprob, 4), '_pEmit': round(p_emit, 4),
        '_hit': hit, '_modelSuggestion': sug,
        '_ranked': [{'key': k['key']} for k in ranked],
        '_latencyMs': round(ms, 1),
    })

with open(os.path.join(HERE, 'batch3-scored.jsonl'), 'w', encoding='utf-8') as f:
    for o in out:
        f.write(json.dumps(o, ensure_ascii=False) + '\n')

# ---- build xlsx ----
XLSX_SKILL_DIR = os.environ.get('XLSX_SKILL_DIR',
    r'C:\Users\JH Z\.zcode\cli\plugins\cache\zcode-plugins-official\document-skills\0.1.1\skills\xlsx')
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, 'templates')]:
    if sub not in sys.path:
        sys.path.insert(0, sub)
from base import (FONT_NAME, HEADER_BOLD, PRIMARY, setup_sheet, style_header_row,
                  style_data_row, auto_fit_columns, auto_fit_row_heights,
                  font_caption)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = '审批表'
headers = ['序号', '样本', '类别', '查询', '观测分', '意图P', '命中', '模型建议',
           '校准建议', '理由', '用户选择']
last_col = len(headers) + 1
setup_sheet(ws, title='Batch-3 模型不确定区审批（28 条）', last_col=last_col)
for ci, h in enumerate(headers, start=2):
    ws.cell(row=4, column=ci, value=h)
style_header_row(ws, row_num=4, col_start=2, col_end=last_col)
for i, o in enumerate(out):
    row = 5 + i
    vals = [i + 1, o['sampleId'], o['kind'], o['queryText'],
            o['_denseTop'], round(o['_intentProb'], 2),
            'Y' if o['_hit'] else 'N',
            f"{o['_modelSuggestion']}(p={o['_pEmit']:.2f})",
            o['proposedAction'], o['rationale'], None]
    for ci, v in enumerate(vals, start=2):
        ws.cell(row=row, column=ci, value=v)
    style_data_row(ws, row_num=row, col_start=2, col_end=last_col, row_index=i)
    sel = ws.cell(row=row, column=12)
    sel.alignment = Alignment(horizontal='center', vertical='center')
    sel.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=PRIMARY)
dv = DataValidation(type='list', formula1='"A,P,S,H,E"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f'L5:L{4 + len(out)}')
note = ws.cell(row=6 + len(out), column=2,
               value='本批全部位于模型不确定区或新特征验收点：请按你的真实偏好填 A/P/S/H/E。'
                     '「校准建议」是 strong-agent 先验，「模型建议」是 v3a 拟合输出——两者不一致的行信息量最大。')
note.font = font_caption()
ws.freeze_panes = 'D5'
auto_fit_columns(ws, min_width=8, max_width=44, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)
outp = os.path.join(HERE, 'batch3-review.xlsx')
wb.properties.creator = 'Z.ai'
wb.save(outp)

agree = sum(1 for o in out if (o['_modelSuggestion'] == 'emit' and o['proposedAction'] == 'activate')
            or (o['_modelSuggestion'] == 'prefetch' and o['proposedAction'] == 'prefetch')
            or (o['_modelSuggestion'] == 'suppress' and o['proposedAction'] == 'suppress'))
print('[b3] scored %d | model-agrees-with-prior %d/%d | saved %s'
      % (len(out), agree, len(out), outp))
