# -*- coding: utf-8 -*-
"""Apply cf-review.xlsx decisions -> gold-confirmed-cf.jsonl (+ merged summary)."""
import json
import os

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
wb = load_workbook(os.path.join(HERE, 'cf-review.xlsx'))
ws = wb['审批表']
cf = {json.loads(l)['sampleId']: json.loads(l) for l in
      open(os.path.join(HERE, 'counterfactual-pairs.jsonl'), encoding='utf-8')
      if l.strip()}


def parse_choice(raw):
    if raw is None:
        return None, None, True
    s = str(raw).strip()
    if '(' in s or '（' in s:
        head = s.replace('（', '(').split('(')[0].strip()
        inner = s[s.find('(') + 1:s.rfind(')')] if ')' in s else ''
    else:
        head, inner = s, ''
    if head.lower() in ('a', 'p', 's', 'h', 'e'):
        return head.upper(), inner or None, False
    return None, s, True


out, log = [], []
for row in range(5, 37):
    sid = ws.cell(row=row, column=3).value
    if not sid:
        continue
    letter, note, deferred = parse_choice(ws.cell(row=row, column=10).value)
    base = cf[sid]
    comments = []
    for c in range(2, 15):
        com = ws.cell(row=row, column=c).comment
        if com is not None:
            comments.append({'cell': ws.cell(row=row, column=c).coordinate,
                             'text': com.text})
    rec = {
        'sampleId': sid, 'pairId': base['pairId'], 'category': base['category'],
        'queryText': base['queryText'], 'language': base.get('language'),
        'rawChoice': str(ws.cell(row=row, column=10).value),
        'choiceLetter': letter, 'choiceNote': note, 'rowComments': comments,
        'previousProposal': {'action': base['proposedAction'],
                             'harmful': bool(base.get('harmful'))},
        'finalExpectedMemoryIds': list(base.get('expectedMemoryIds') or []),
        'finalForbiddenMemoryIds': list(base.get('forbiddenMemoryIds') or []),
        'labelSource': 'human', 'isGold': False,
    }
    if deferred:
        rec['status'] = 'deferred'
        deferred_log = True
        out.append(rec)
        log.append((sid, 'DEFERRED'))
        continue
    rec['isGold'] = True
    rec['finalAction'] = letter
    rec['finalHarmful'] = bool(base.get('harmful')) and letter == 'S' \
        if False else bool(base.get('harmful'))
    if letter != base['proposedAction'].upper()[0]:
        rec['overridesProposal'] = True
    out.append(rec)
    log.append((sid, letter, 'OVERRIDE' if rec.get('overridesProposal') else ''))

with open(os.path.join(HERE, 'gold-confirmed-cf.jsonl'), 'w', encoding='utf-8') as f:
    for r in out:
        f.write(json.dumps(r, ensure_ascii=False) + '\n')

from collections import Counter
cnt = Counter(r['finalAction'] for r in out if r.get('isGold'))

# merge totals with batch 1
b1 = [json.loads(l) for l in open(os.path.join(HERE, 'gold-confirmed.jsonl'),
                                  encoding='utf-8') if l.strip()]
c1 = Counter(r['finalAction'] for r in b1 if r.get('isGold'))
total = {k: c1.get(k, 0) + cnt.get(k, 0) for k in ('A', 'P', 'S')}
summary = {
    'batch2': {'gold': sum(cnt.values()), 'byAction': dict(cnt),
               'overrides': [sid for sid, l, o in log if o]},
    'combinedGold': {'byAction': total, 'total': sum(total.values()),
                     'meetsQuota15PerClass':
                         all(v >= 15 for v in total.values())},
    'note': 'cf-008/cf-014 由 suppress 改判 P：跨项目可顺便介绍（用户联想哲学，与议题③a一致）',
}
with open(os.path.join(HERE, 'user-decisions-cf-applied.json'), 'w',
          encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=1)
print(json.dumps(summary, ensure_ascii=False, indent=1))
