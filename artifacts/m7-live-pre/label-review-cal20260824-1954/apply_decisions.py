# -*- coding: utf-8 -*-
"""Apply user decisions from boundary-review.xlsx -> gold-confirmed.jsonl +
user-decisions-applied.json. Faithful transcription: annotated choices kept
verbatim; deferred/settings-gated rows NOT flipped to gold."""
import json
import os

from openpyxl import load_workbook

HERE = os.path.dirname(os.path.abspath(__file__))
wb = load_workbook(os.path.join(HERE, 'boundary-review.xlsx'))
ws = wb['审批表']

reviewed = {}
for line in open(os.path.join(HERE, 'existing-labels-reviewed.jsonl'),
                 encoding='utf-8'):
    if line.strip():
        o = json.loads(line)
        reviewed[o['sampleId']] = o


def parse_choice(raw):
    """Return (letter|None, note|None, deferred)."""
    if raw is None:
        return None, None, True
    s = str(raw).strip()
    inner = ''
    if '(' in s or '（' in s:
        head = s.replace('（', '(').split('(')[0].strip()
        inner = s[s.find('(') + 1: s.rfind(')')] if ')' in s else ''
    else:
        head = s
    head_l = head.strip().lower()
    if head_l in ('a', 'p', 's', 'h', 'e'):
        return head_l.upper(), (inner or None), False
    if not head_l and s:
        return None, s, True          # pure-comment cell = deferred to settings
    return None, s, True              # unparsable -> deferred, keep verbatim


gold, deferred, log = [], [], []
for row in range(5, 33):
    sid = ws.cell(row=row, column=3).value
    if not sid:
        continue
    raw_choice = ws.cell(row=row, column=10).value
    # collect any comments anywhere in the row
    comments = []
    for c in range(2, 15):
        com = ws.cell(row=row, column=c).comment
        if com is not None:
            comments.append({'cell': ws.cell(row=row, column=c).coordinate,
                             'text': com.text})
    letter, note, deferred_flag = parse_choice(raw_choice)
    base = reviewed[sid]
    rec = {
        'sampleId': sid,
        'queryText': base['queryText'],
        'surface': base['surface'],
        'rawChoice': str(raw_choice) if raw_choice is not None else None,
        'choiceLetter': letter,
        'choiceNote': note,
        'rowComments': comments,
        'previousProposal': {'action': base['proposedAction'],
                             'harmful': base['harmful'],
                             'expectedMemoryIds': base['proposedExpectedMemoryIds'],
                             'forbiddenMemoryIds': base['proposedForbiddenMemoryIds']},
        'labelSource': 'human',
        'isGold': False,
    }
    if deferred_flag:
        rec['status'] = 'deferred-to-user-setting'
        rec['settingTopic'] = 'personalization-③b 敏感度分档（直接提醒/仅搜索可用/主动提醒）'
        deferred.append(rec)
        log.append((sid, 'DEFERRED', (note or '')[:60]))
        continue
    rec['isGold'] = True
    rec['finalAction'] = letter
    if sid == 'cal-0058':
        rec['finalHarmful'] = False   # user overrode: advisory cross-ws relay
    elif sid == 'cal-0060':
        rec['finalHarmful'] = False   # user: advisory reference, AI decides
    elif sid == 'cal-0059':
        rec['finalHarmful'] = False   # user: explicit user authority governs
    else:
        rec['finalHarmful'] = base['harmful']
    exp = list(base['proposedExpectedMemoryIds'])
    forb = list(base['proposedForbiddenMemoryIds'])
    if sid == 'cal-0036':
        exp, forb = ['ep_d1ad532209bc0390'], []
    if sid == 'cal-0037':
        exp, forb = ['ep_fed40ce72e0ecc0f'], []
    if sid in ('cal-0055',):
        exp, forb = ['ep_e515177f4c632f2c'], []
    if sid == 'cal-0058':
        exp, forb = ['ep_e515177f4c632f2c'], ['ep_fc43a607a1cf33a1']
    rec['finalExpectedMemoryIds'] = exp
    rec['finalForbiddenMemoryIds'] = forb
    if sid in ('cal-0036', 'cal-0037', 'cal-0055', 'cal-0058'):
        rec['requiresCrossWorkspaceRelay'] = True
        rec['relayNote'] = '目标在外部工作区：当前三重过滤下不可达，gold 生效以未来跨工作区联想设置(议题③a)为前提'
    if letter != base['proposedAction'].upper()[0] and \
            letter != ('H' if base['harmful'] else base['proposedAction'].upper()[0]):
        rec['overridesProposal'] = True
    gold.append(rec)
    log.append((sid, letter, (note or '')[:50]))

with open(os.path.join(HERE, 'gold-confirmed.jsonl'), 'w', encoding='utf-8') as f:
    for rec in gold:
        f.write(json.dumps(rec, ensure_ascii=False) + '\n')

from collections import Counter
cnt = Counter(r['finalAction'] for r in gold)
summary = {
    'appliedAt': '2026-08-24',
    'source': 'boundary-review.xlsx 审批表',
    'goldCount': len(gold),
    'goldByAction': dict(cnt),
    'deferredToSetting': [r['sampleId'] for r in deferred],
    'overridesOfAgentProposal': [
        {'sampleId': r['sampleId'], 'proposal': r['previousProposal']['action'],
         'user': r['finalAction'], 'note': r.get('choiceNote')}
        for r in gold if r.get('overridesProposal')],
    'crossWorkspaceRelayGolds': [r['sampleId'] for r in gold
                                 if r.get('requiresCrossWorkspaceRelay')],
    'quotaStatus': {'activate': cnt.get('A', 0), 'prefetch': cnt.get('P', 0),
                    'suppress': cnt.get('S', 0),
                    'targetPerClass': 15,
                    'gap': {'activate': max(0, 15 - cnt.get('A', 0)),
                            'prefetch': max(0, 15 - cnt.get('P', 0)),
                            'suppress': max(0, 15 - cnt.get('S', 0))}},
}
with open(os.path.join(HERE, 'user-decisions-applied.json'), 'w',
          encoding='utf-8') as f:
    json.dump(summary, f, ensure_ascii=False, indent=1)
print(json.dumps(summary, ensure_ascii=False, indent=1))
for sid, letter, note in log:
    print(sid, letter, '|', note)
