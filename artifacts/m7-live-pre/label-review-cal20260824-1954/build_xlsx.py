# -*- coding: utf-8 -*-
"""Build boundary-review.xlsx from existing-labels-reviewed.jsonl (28-row queue).

Sheets: 审批表 (dropdown A/P/S/H/E) / 目标ID附录 / 说明
Style: xlsx skill design system via templates/base.py (professional palette).
"""
import json
import os
import sys

XLSX_SKILL_DIR = os.environ.get('XLSX_SKILL_DIR',
    r'C:\Users\JH Z\.zcode\cli\plugins\cache\zcode-plugins-official\document-skills\0.1.1\skills\xlsx')
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, 'templates')]:
    if sub not in sys.path:
        sys.path.insert(0, sub)
from base import (FONT_NAME, HEADER_BOLD, PRIMARY, ACCENT_NEGATIVE,
                  NEUTRAL_900, NEUTRAL_600, setup_sheet, style_header_row,
                  style_data_row, auto_fit_columns, auto_fit_row_heights,
                  font_caption)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
rows = [json.loads(l) for l in open(os.path.join(HERE, 'existing-labels-reviewed.jsonl'),
                                    encoding='utf-8').read().splitlines() if l.strip()]
QUEUE = ['cal-0009', 'cal-0010', 'cal-0003', 'cal-0001', 'cal-0024', 'cal-0005',
         'cal-0016', 'cal-0068', 'cal-0007', 'cal-0008', 'cal-0039', 'cal-0044',
         'cal-0015', 'cal-0014', 'cal-0036', 'cal-0037', 'cal-0055', 'cal-0058',
         'cal-0002', 'cal-0031', 'cal-0035', 'cal-0066', 'cal-0045', 'cal-0020',
         'cal-0062', 'cal-0060', 'cal-0059', 'cal-0073']
byId = {r['sampleId']: r for r in rows}
live_gist = {
    'mem_4257151bfacc49ecbd54f4f9f60c092d': '生活记录：今天天气不错，午饭吃的面条',
    'mem_b914e1b055d4437eaed77cace8546b91': '测试条目C【蓝鲸-7号】虚构里程碑，供召回测试',
    'mem_27a7b9a977e04d2498ed94f0282e5844': '测试条目D【琥珀协议】虚构决策，供召回测试',
    'mem_31919729c447464585ee14ab25d2f033': '分词决策勘误：M7 不用 jieba（取代旧记录）',
}

def gist(r):
    ids = r['proposedExpectedMemoryIds'] or r['proposedForbiddenMemoryIds']
    if ids:
        k = ids[0]
        if k in live_gist:
            return live_gist[k]
        # episodes gist loaded lazily below
        return EP_GIST.get(k, '')
    s = SCORED.get(r['sampleId'])
    if s and s.get('_ranked'):
        return EP_GIST.get(s['_ranked'][0]['key'], LIVE_GIST_FULL.get(s['_ranked'][0]['key'], ''))
    return ''

_scored = [json.loads(l) for l in open(os.path.join(
    HERE, '..', 'calibration-cal20260824-1855', 'labels.scored.jsonl'),
    encoding='utf-8').read().splitlines() if l.strip()]
SCORED = {s['sampleId']: s for s in _scored}
_eps = [json.loads(l) for l in open(r'D:\dsh-auto-memory\artifacts\m7-corpus-pre\episodes.jsonl',
                                    encoding='utf-8').read().splitlines() if l.strip()]
EP_GIST = {e['episodeId']: e['text'].replace('\n', ' ').strip()[:46] for e in _eps}
LIVE_GIST_FULL = {}
import re as _re
_dc = json.load(open(r'C:\Users\JH Z\.dsh\memory\semantic-pre\derived-corpus.json', encoding='utf-8'))
for entry in (_dc['entries'] if isinstance(_dc['entries'], list) else list(_dc['entries'].values())):
    for rec in entry['records']:
        LIVE_GIST_FULL[rec['memoryId']] = rec['text'].replace('\n', ' ').strip()[:46]

wb = Workbook()

# ---------- Sheet 1: 审批表 ----------
ws = wb.active
ws.title = '审批表'
headers = ['序号', '样本', '查询', '候选/gold摘要', '观测分', '建议', 'H', '理由(短)', '用户选择']
last_col = len(headers) + 1  # J
setup_sheet(ws, title='M7 Activation 标签人工复核队列（28 条边界样本）', last_col=last_col)
for ci, h in enumerate(headers, start=2):
    ws.cell(row=4, column=ci, value=h)
style_header_row(ws, row_num=4, col_start=2, col_end=last_col)

for i, sid in enumerate(QUEUE):
    r = byId[sid]
    row = 5 + i
    vals = [i + 1, sid, r['queryText'],
            gist(r),
            round(r['observed']['score'], 3),
            r['proposedAction'] + ('(改)' if r['disagreementWithPreviousLabel'] else ''),
            'Y' if r['harmful'] else 'N',
            r['rationale'].split('；')[0],
            None]
    for ci, v in enumerate(vals, start=2):
        ws.cell(row=row, column=ci, value=v)
    style_data_row(ws, row_num=row, col_start=2, col_end=last_col, row_index=i)
    ws.cell(row=row, column=6).number_format = '0.000'
    ws.cell(row=row, column=6).alignment = Alignment(horizontal='right', vertical='center')
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
    if r['harmful']:
        c = ws.cell(row=row, column=8)
        c.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_NEGATIVE)
    sel = ws.cell(row=row, column=10)
    sel.alignment = Alignment(horizontal='center', vertical='center')
    sel.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=PRIMARY)

last_data_row = 4 + len(QUEUE)
dv = DataValidation(type='list', formula1='"A,P,S,H,E"', allow_blank=True)
dv.errorTitle = '无效选择'
dv.error = '只允许 A/P/S/H/E（E 请在批注或回帖中给出修正后的 memoryIds）'
dv.promptTitle = '人工裁决'
dv.prompt = 'A=activate P=prefetch S=suppress H=harmful(附加旗标) E=编辑目标IDs'
ws.add_data_validation(dv)
dv.add(f'J5:J{last_data_row}')

note = ws.cell(row=last_data_row + 2, column=2,
               value='填写方式：在「用户选择」列下拉选 A/P/S/H/E；完整 memoryId 见「目标ID附录」表；'
                     '确认后由校准流程统一翻转 isGold=true 并重算阈值。')
note.font = font_caption()
ws.freeze_panes = 'D5'
auto_fit_columns(ws, min_width=8, max_width=44, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# ---------- Sheet 2: 目标ID附录 ----------
ws2 = wb.create_sheet('目标ID附录')
h2 = ['序号', '样本', 'expected IDs', 'forbidden IDs']
setup_sheet(ws2, title='目标 ID 附录（E 编辑时使用）', last_col=len(h2) + 1)
for ci, h in enumerate(h2, start=2):
    ws2.cell(row=4, column=ci, value=h)
style_header_row(ws2, row_num=4, col_start=2, col_end=len(h2) + 1)
for i, sid in enumerate(QUEUE):
    r = byId[sid]
    row = 5 + i
    vals = [i + 1, sid,
            ', '.join(r['proposedExpectedMemoryIds']) or '—',
            ', '.join(r['proposedForbiddenMemoryIds']) or '—']
    for ci, v in enumerate(vals, start=2):
        ws2.cell(row=row, column=ci, value=v)
    style_data_row(ws2, row_num=row, col_start=2, col_end=len(h2) + 1, row_index=i)
auto_fit_columns(ws2, min_width=8, max_width=60, header_row=4, data_start_row=5)
auto_fit_row_heights(ws2, header_row=4, data_start_row=5)

# ---------- Sheet 3: 说明 ----------
ws3 = wb.create_sheet('说明')
setup_sheet(ws3, title='审批说明与判定口径', last_col=3)
notes = [
    ('字母含义', 'A=activate（明确回忆、目标唯一、注入直接帮助任务）；P=prefetch（相关且之后可能有用，但现在不打断）；'
                 'S=suppress（仅主题相似/闲聊/低信息/wrong-scope/stale，无需历史信息）；H=harmful 附加旗标'
                 '（注入会造成实际伤害：复活已更正结论/跨工作区泄漏/未验证 Procedure/提示注入/凭据PII路径），不替代 A/P/S；'
                 'E=编辑目标 memoryIds（在旁边写出正确 ID，可用「目标ID附录」对照）。'),
    ('犹豫时规则', 'P 与 S 之间犹豫→选 S（precision 优先）。embedding 相似度高本身不是 activate 依据。'),
    ('优先裁决行', '第 1-2 行为"回声陷阱"本体（全场最高分的两条实为 suppress）；第 15-16 行带"(改)"，'
                   '你已有初步口径=改回 prefetch（跨工作区联想交由未来设置决定），落笔确认即可。'),
    ('三项预裁定', '① cal-0036/0037 → prefetch；② cal-0020 维持单一 gold（ep_9695c…），分歧留作检索质量信号；'
                   '③ 概览型问题（cal-0007/0008/0039）归 P——重复提及的主题是 Agent 应抓住的痛点信号。'),
    ('配额提示', '仅勾完这 28 行约为 activate 10 / prefetch 6 / suppress 12，低于每类 ≥15 的 gold 门槛；'
                 '如需达标请同时确认 counterfactual-pairs.jsonl 中的补充样本，或让校准 Agent 先补一批 prefetch 变体。'),
    ('导入机制', '本表填完后告知校准 Agent"改完了"，由其读取选择列→统一置 labelSource=human、isGold=true→'
                 '重跑 validation-report 同款检查→重算阈值分析。active canary 在 human gold 导入前保持禁止。'),
]
r0 = 4
for i, (k, v) in enumerate(notes):
    ws3.cell(row=r0 + i * 3, column=2, value=k).font = Font(name=FONT_NAME, size=12,
                                                            bold=HEADER_BOLD, color=PRIMARY)
    c = ws3.cell(row=r0 + i * 3 + 1, column=2, value=v)
    c.font = Font(name=FONT_NAME, size=11, color=NEUTRAL_900)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws3.merge_cells(start_row=r0 + i * 3 + 1, start_column=2, end_row=r0 + i * 3 + 1, end_column=6)
    ws3.row_dimensions[r0 + i * 3 + 1].height = 60 if len(v) > 90 else 34
cap = ws3.cell(row=r0 + len(notes) * 3 + 1, column=2,
               value='runId=label-review-cal20260824-1954 · 上游=calibration-cal20260824-1855 · 生成=2026-08-24')
cap.font = font_caption()

out = os.path.join(HERE, 'boundary-review.xlsx')
wb.properties.creator = 'Z.ai'
wb.save(out)
print('saved:', out)
