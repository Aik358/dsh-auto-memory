# -*- coding: utf-8 -*-
"""Build cf-review.xlsx: pre-screened counterfactual samples for the second
quick confirmation pass (fills gold quota gaps: S+11 A+4 P+4)."""
import json
import os
import sys

XLSX_SKILL_DIR = os.environ.get('XLSX_SKILL_DIR',
    r'C:\Users\JH Z\.zcode\cli\plugins\cache\zcode-plugins-official\document-skills\0.1.1\skills\xlsx')
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, 'templates')]:
    if sub not in sys.path:
        sys.path.insert(0, sub)
from base import (FONT_NAME, HEADER_BOLD, PRIMARY, ACCENT_NEGATIVE,
                  setup_sheet, style_header_row, style_data_row,
                  auto_fit_columns, auto_fit_row_heights, font_caption)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
cf = [json.loads(l) for l in open(os.path.join(HERE, 'counterfactual-pairs.jsonl'),
                                  encoding='utf-8').read().splitlines() if l.strip()]
groups = {}
order = []
for s in cf:
    groups.setdefault(s['pairId'], []).append(s)
    if s['pairId'] not in order:
        order.append(s['pairId'])

def pick(pair, idx):          # idx 1-based within pair
    return groups[pair][idx - 1]

# (pairId, idx) selections: fill S first, then A, then P
SEL = [
    ('cf-g01', 2), ('cf-g02', 2), ('cf-g03', 2), ('cf-g04', 2), ('cf-g06', 2),
    ('cf-g07', 2), ('cf-g08', 2), ('cf-g09', 2), ('cf-g10', 2),
    ('cf-g51', 1), ('cf-g52', 1), ('cf-g53', 1), ('cf-g54', 1), ('cf-g55', 1), ('cf-g56', 1),
    ('cf-g01', 1), ('cf-g02', 1), ('cf-g03', 1), ('cf-g04', 1), ('cf-g11', 1),
    ('cf-g12', 1), ('cf-g13', 1), ('cf-g20', 1), ('cf-g45', 1), ('cf-g48', 1), ('cf-g49', 1),
    ('cf-g04', 3), ('cf-g05', 3), ('cf-g13', 2), ('cf-g17', 2), ('cf-g18', 2), ('cf-g20', 2),
]
rows = [pick(p, i) for p, i in SEL]

_eps = [json.loads(l) for l in open(r'D:\dsh-auto-memory\artifacts\m7-corpus-pre\episodes.jsonl',
                                    encoding='utf-8').read().splitlines() if l.strip()]
EP_GIST = {e['episodeId']: e['text'].replace('\n', ' ').strip()[:46] for e in _eps}
_dc = json.load(open(r'C:\Users\JH Z\.dsh\memory\semantic-pre\derived-corpus.json',
                     encoding='utf-8'))
LIVE_GIST = {}
for entry in (_dc['entries'] if isinstance(_dc['entries'], list)
              else list(_dc['entries'].values())):
    for rec in entry['records']:
        LIVE_GIST[rec['memoryId']] = rec['text'].replace('\n', ' ').strip()[:46]

def gist(s):
    pid = s.get('parentMemoryId') or s.get('parentEpisodeId')
    return LIVE_GIST.get(pid) or EP_GIST.get(pid) or ''

wb = Workbook()
ws = wb.active
ws.title = '审批表'
headers = ['序号', '样本', '类别', '查询', '候选/gold摘要', '建议', 'H', '理由(短)', '用户选择']
last_col = len(headers) + 1
setup_sheet(ws, title='Counterfactual 补充确认队列（34 条，填 gold 配额缺口）', last_col=last_col)
for ci, h in enumerate(headers, start=2):
    ws.cell(row=4, column=ci, value=h)
style_header_row(ws, row_num=4, col_start=2, col_end=last_col)
for i, s in enumerate(rows):
    row = 5 + i
    vals = [i + 1, s['sampleId'], s['category'], s['queryText'], gist(s),
            s['proposedAction'] + ('(改)' if False else ''), 'Y' if s.get('harmful') else 'N',
            (s.get('rationale') or '').split('；')[0], None]
    for ci, v in enumerate(vals, start=2):
        ws.cell(row=row, column=ci, value=v)
    style_data_row(ws, row_num=row, col_start=2, col_end=last_col, row_index=i)
    ws.cell(row=row, column=8).alignment = Alignment(horizontal='center', vertical='center')
    if s.get('harmful'):
        c = ws.cell(row=row, column=8)
        c.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_NEGATIVE)
    sel = ws.cell(row=row, column=10)
    sel.alignment = Alignment(horizontal='center', vertical='center')
    sel.font = Font(name=FONT_NAME, size=12, bold=HEADER_BOLD, color=PRIMARY)
last_data_row = 4 + len(rows)
dv = DataValidation(type='list', formula1='"A,P,S,H,E"', allow_blank=True)
dv.errorTitle = '无效选择'
dv.error = '只允许 A/P/S/H/E'
ws.add_data_validation(dv)
dv.add(f'J5:J{last_data_row}')
note = ws.cell(row=last_data_row + 2, column=2,
               value='这批是同一真实记忆在不同对话意图下的最小对照：B 面多为"内容相似但只是陈述/闲聊"→S，'
                     'A 面为显式回忆→activate，C 面为任务语境备用→prefetch。犹豫时 P/S 之间选 S。'
                     '全部填完说"改完了"，与第一批 gold 合并后重算阈值。')
note.font = font_caption()
ws.freeze_panes = 'D5'
auto_fit_columns(ws, min_width=8, max_width=44, header_row=4, data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

out = os.path.join(HERE, 'cf-review.xlsx')
wb.properties.creator = 'Z.ai'
wb.save(out)
from collections import Counter
print('saved:', out, '| rows:', len(rows),
      '| actions:', dict(Counter(s['proposedAction'] for s in rows)))
