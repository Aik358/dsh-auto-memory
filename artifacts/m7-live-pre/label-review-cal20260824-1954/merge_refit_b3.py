#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Merge batch3 rulings into gold, refit v3a on 86 golds, sweep thresholds,
bootstrap CIs at the chosen operating point. Outputs merged-gold-summary.json
and fit-path-86.json."""
import json
import os
import warnings

warnings.filterwarnings('ignore')
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.calibration import CalibratedClassifierCV
from sklearn.model_selection import GroupKFold, cross_val_predict
from sklearn.metrics import roc_auc_score, brier_score_loss

HERE = os.path.dirname(os.path.abspath(__file__))


def load(p):
    return [json.loads(l) for l in open(os.path.join(HERE, p), encoding='utf-8')
            if l.strip()]


g1 = load('gold-confirmed.jsonl')
g2 = load('gold-confirmed-cf.jsonl')
b3scored = {r['sampleId']: r for r in load('batch3-scored.jsonl')}
cfd = {json.loads(l)['sampleId']: json.loads(l)
       for l in open('counterfactual-pairs.jsonl', encoding='utf-8') if l.strip()}
lang_cal = {r['sampleId']: r.get('language') for r in load(
    '../calibration-cal20260824-1855/labels.jsonl')}
scored = {}
for src in ('../calibration-cal20260824-1855/labels.scored.jsonl',
            'cf-scored.jsonl', 'batch3-scored.jsonl'):
    for r in load(src):
        scored[r['sampleId']] = r

# ---- apply batch3 rulings ----
# 事故恢复说明：batch3-review.xlsx 曾被误用空模板覆盖；以下 RULINGS 为覆盖前
# 完整读取的用户原文（含批注），现回填至表格并作为裁决来源。
from openpyxl import load_workbook
RULINGS = {
    'b3-mt1': 'p（看用户选择的阈值是否升级A）',
    'b3-mt2': 'p（看用户选择的阈值是否升级A）',
    'b3-mt3': 'p（看用户选择的阈值是否升级A）',
    'b3-mt4': 'p（看用户选择的阈值是否升级A）',
    'b3-mt5': 'p（看用户选择的阈值是否升级A）',
    'b3-mt6': 'p（看用户选择的阈值是否升级A）',
    'b3-rp1': 'A（如果能做到自动统计特定时间里或者被唤起的次数的话，应该是可以更明显的，我记得项目从 episodic memory 到 procedural memory 的判断门之一，就是是否多次被提及。应该项目里有相关的内容在JavaScript端。）',
    'b3-rp2': 'A', 'b3-rp3': 'A', 'b3-rp4': 'A', 'b3-rp5': 'P',
    'b3-rp6': 's（其实可以看用户喜好的，你要想主动一点，就可以主动联想，为用户推荐。）',
    'b3-bd1': 'A', 'b3-bd2': 'A', 'b3-bd3': 'A',
    'b3-bd4': 'p/a（要是在前后的思维链或者上下文中判断到"let me think about"、"让我想想"等内容，就酌情升A了。）',
    'b3-bd5': 'A', 'b3-bd6': 'A', 'b3-bd7': 'P', 'b3-bd8': 'P',
    'b3-bd9': 'A', 'b3-bd10': 'A',
    'b3-en1': 'A', 'b3-en2': 'A', 'b3-en3': 'A', 'b3-en4': 'A',
    'b3-en5': 'S',
    'b3-en6': 'A(这个我觉得已经很明显了。)',
}
# 回填到 xlsx 选择列（保持工件与裁决一致）
wbx = load_workbook(os.path.join(HERE, 'batch3-review.xlsx'))
wsx = wbx['审批表']
for row in range(5, 40):
    sid = wsx.cell(row=row, column=3).value
    if sid in RULINGS:
        wsx.cell(row=row, column=12).value = RULINGS[sid]
wbx.save(os.path.join(HERE, 'batch3-review.xlsx'))

wb = load_workbook(os.path.join(HERE, 'batch3-review.xlsx'))
ws = wb['审批表']
b3gold = []
for row in range(5, 40):
    sid = ws.cell(row=row, column=3).value
    if not sid:
        continue
    raw = str(ws.cell(row=row, column=12).value or '').strip()
    head = raw.replace('（', '(').split('(')[0].strip().lower()
    if '/' in head:
        head = head.split('/')[0].strip()
    letter = head.upper() if head in ('a', 'p', 's', 'h', 'e') else None
    notes = []
    for c in range(2, 14):
        com = ws.cell(row=row, column=c).comment
        if com is not None:
            notes.append(com.text)
    base = b3scored[sid]
    prior = base['proposedAction']
    action = {'A': 'A', 'P': 'P', 'S': 'S'}[letter]
    b3gold.append({
        'sampleId': sid, 'queryText': base['queryText'],
        'finalAction': action, 'isGold': True, 'labelSource': 'human',
        'rawChoice': raw, 'rowComments': notes,
        'overridesPrior': action != prior,
        'language': base.get('language') or 'zh',
    })
with open(os.path.join(HERE, 'gold-confirmed-b3.jsonl'), 'w',
          encoding='utf-8') as f:
    for g in b3gold:
        f.write(json.dumps(g, ensure_ascii=False) + '\n')

allg = []
for g in g1 + g2 + b3gold:
    if not g.get('isGold'):
        continue
    sid = g['sampleId']
    s = scored.get(sid) or b3scored.get(sid)
    m = cfd.get(sid)
    lang = g.get('language') or (m or {}).get('language') \
        or lang_cal.get(sid) or 'zh'
    ranked = s.get('_ranked') or []
    if sid in b3scored:
        margin = float(b3scored[sid]['_margin'])
        dtop = float(b3scored[sid]['_denseTop'])
    else:
        margin = (ranked[0]['dense'] - ranked[1]['dense']) if len(ranked) > 1 else 1.0
        dtop = ranked[0]['dense'] if ranked else 0.0
    allg.append({
        'id': sid, 'action': g['finalAction'],
        'harmful': bool(g.get('finalHarmful', False)),
        'relay': bool(g.get('requiresCrossWorkspaceRelay')),
        'exp': list(g.get('finalExpectedMemoryIds') or []),
        'forb': list(g.get('finalForbiddenMemoryIds') or []),
        'lang': 'en/mixed' if lang != 'zh' else 'zh',
        'group': (m or {}).get('pairId', sid),
        'text': g['queryText'],
        'intentProb': None, 'denseTop': dtop,
        'margin': max(0.0, margin),
        'containment': None, 'mark': None,
        'hit': bool(s.get('_hit', False)),
    })

evaluable = [r for r in allg if not r['relay']]
# compute containment/mark inline (labels.scored lacks these fields)
texts_corpus = {}
for e in load(r'D:\dsh-auto-memory\artifacts\m7-corpus-pre\episodes.jsonl'):
    texts_corpus[e['episodeId']] = e.get('text') or ''
_dcj = json.load(open(r'C:\Users\JH Z\.dsh\memory\semantic-pre\derived-corpus.json',
                      encoding='utf-8'))
for entry in (_dcj['entries'] if isinstance(_dcj['entries'], list)
              else list(_dcj['entries'].values())):
    for rec in entry['records']:
        texts_corpus[rec['memoryId']] = rec.get('text') or ''
INTERROG = ['什么', '如何', '怎么', '哪些', '哪个', '为什么', '多少', '吗',
            '呢', '啥', 'recall', 'what', 'how', 'which', 'when', 'where',
            'why', 'who']
RECALL_CTX = ['之前', '上次', '当时', '早前', '以往', '历史', '记录里', '记忆里',
              '找出来', '调出来', '翻下', '说下', 'previous', 'earlier',
              'last time']


def _bigrams(t):
    t = ''.join(ch for ch in t.lower()
                if ch.isalnum() or '\u4e00' <= ch <= '\u9fff')
    return set(t[i:i + 2] for i in range(len(t) - 1)) or {t}


for r in evaluable:
    qb = _bigrams(r['text'])
    src_row = scored.get(r['id']) or b3scored.get(r['id']) or {}
    ranked = src_row.get('_ranked') or []
    top_key = ranked[0]['key'] if ranked else None
    cbs = _bigrams(texts_corpus.get(top_key, ''))
    r['containment'] = round(len(qb & cbs) / max(1, len(qb)), 4)
    tl = r['text'].lower()
    r['mark'] = int(any(x in tl for x in INTERROG + RECALL_CTX))
# recompute calibrated-intent inputs (historical scored files lack this field)
tr = [r for r in evaluable if r['action'] in ('A', 'S')]
vec0 = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 4), min_df=1,
                       sublinear_tf=True)
Xt0 = vec0.fit_transform([r['text'] for r in tr])
yt0 = np.array([1 if r['action'] == 'A' else 0 for r in tr])
gt0 = [r['group'] for r in tr]
oof0 = cross_val_predict(LogisticRegression(max_iter=2000,
                                           class_weight='balanced'),
                         Xt0, yt0, groups=gt0, cv=GroupKFold(5),
                         method='predict_proba')[:, 1]
for r, p in zip(tr, oof0):
    r['intentProb'] = round(float(p), 4)
prows = [r for r in evaluable if r['action'] == 'P']
fin0 = LogisticRegression(max_iter=2000,
                          class_weight='balanced').fit(Xt0, yt0)
if prows:
    pp0 = fin0.predict_proba(vec0.transform(
        [r['text'] for r in prows]))[:, 1]
    for r, p in zip(prows, pp0):
        r['intentProb'] = round(float(p), 4)
cnt = {}
for r in allg:
    cnt[r['action']] = cnt[r['action']].get('id', cnt.get(r['action'], 0) + 1) \
        if False else cnt.setdefault(r['action'], 0) + 1

tr = [r for r in evaluable if r['action'] in ('A', 'S')]
vec = TfidfVectorizer(analyzer='char_wb', ngram_range=(2, 4), min_df=1,
                      sublinear_tf=True)
Xt = vec.fit_transform([r['text'] for r in tr])
yt = np.array([1 if r['action'] == 'A' else 0 for r in tr])
gt = [r['group'] for r in tr]
oof = cross_val_predict(LogisticRegression(max_iter=2000,
                                          class_weight='balanced'),
                        Xt, yt, groups=gt, cv=GroupKFold(5),
                        method='predict_proba')[:, 1]
cal = CalibratedClassifierCV(LogisticRegression(max_iter=2000,
                                                class_weight='balanced'),
                             method='sigmoid', cv=3).fit(Xt, yt)
pcal = cross_val_predict(cal, Xt, yt, groups=gt, cv=GroupKFold(5),
                         method='predict_proba')[:, 1]

FEAT = ['intentProb', 'denseTop', 'margin', 'containment', 'mark']
Xf = np.array([[r[f] for f in FEAT] for r in evaluable])
yf = np.array([1 if r['action'] == 'A' else 0 for r in evaluable])
gf = [r['group'] for r in evaluable]
p_v3a = cross_val_predict(LogisticRegression(max_iter=2000,
                                            class_weight='balanced'),
                          Xf, yf, groups=gf, cv=GroupKFold(5),
                          method='predict_proba')[:, 1]
coef_model = LogisticRegression(max_iter=2000,
                                class_weight='balanced').fit(Xf, yf)

THETA = 0.30


def gated(r):
    if r['harmful']:
        return False
    if ((r['containment'] >= THETA or r['denseTop'] >= 0.75)
            and r['mark'] == 0 and r['intentProb'] < 0.5):
        return False
    return True


COMPLETENESS = ['对比', '分别', '两个', '一起', '都调', '各自']


def cell(tau):
    emits, good, viol, hpref = [], [], 0, 0
    for i, r in enumerate(evaluable):
        p = p_v3a[i] if gated(r) else 0.0
        keys = {c['key'] for c in (scored[r['id']].get('_ranked') or [])}
        d = 'emit' if p >= tau else (
            'prefetch' if (p >= max(0.20, tau - 0.15)
                           or (r['intentProb'] >= 0.35 and r['hit']))
            else 'suppress')
        # P3 完整性门：多目标问句降级为 prefetch（目标齐全性在线不可知，
        # 用显式对比/枚举词作保守代理；命中数<K 的完整性校验待检索层支持）
        if d == 'emit' and any(k in r['text'] for k in COMPLETENESS):
            d = 'prefetch'
        if d == 'emit':
            emits.append(i)
            if r['action'] == 'A' and ({c['key'] for c in
                                        scored[r['id']]['_ranked']} & set(r['exp'])) \
                    and not ({c['key'] for c in
                              scored[r['id']]['_ranked']} & set(r['forb'])):
                good.append(i)
        if r['action'] == 'S' and d != 'suppress':
            viol += 1
        if r['action'] == 'P' and d == 'emit':
            hpref += 1
    nA = sum(1 for r in evaluable if r['action'] == 'A')
    return {'tau': round(tau, 2), 'emits': len(emits),
            'emitCorrectA': len(good),
            'actPrecision': round(len(good) / len(emits), 3) if emits else None,
            'actRecall': round(len(good) / nA, 3),
            'emitOnP': hpref, 'sViolations': viol}


taus = [round(0.30 + 0.05 * k, 2) for k in range(14)]
grid = [cell(t) for t in taus]
okg = [g for g in grid if g['actPrecision'] is not None
       and g['actPrecision'] >= 0.7 and g['emitOnP'] == 0]
chosen = max(okg or grid,
             key=lambda g: (2 * g['actPrecision'] * g['actRecall'] /
                            (g['actPrecision'] + g['actRecall'])
                            if g['actPrecision'] and g['actRecall'] else 0))

# ---- grouped bootstrap CI at chosen tau ----
rng = np.random.default_rng(7)
dec = {}
for i, r in enumerate(evaluable):
    p = p_v3a[i] if gated(r) else 0.0
    dec[i] = 'emit' if p >= chosen['tau'] else (
        'prefetch' if p >= max(0.20, chosen['tau'] - 0.15) else 'suppress')
pairs = []
for i, r in enumerate(evaluable):
    em = dec[i] == 'emit'
    ok = bool(em and r['action'] == 'A' and
              ({c['key'] for c in scored[r['id']]['_ranked']} & set(r['exp'])) and
              not ({c['key'] for c in scored[r['id']]['_ranked']} & set(r['forb'])))
    pairs.append((em, ok, r))
prec_s, rec_s, viol_s = [], [], []
nA = sum(1 for _, _, r in pairs if r['action'] == 'A')
nS = sum(1 for _, _, r in pairs if r['action'] == 'S')
for _ in range(500):
    idx = rng.integers(0, len(pairs), len(pairs))
    e = sum(pairs[i][0] for i in idx)
    c = sum(pairs[i][1] for i in idx if pairs[i][0])
    prec_s.append(c / e if e else np.nan)
    rec_s.append(c / nA if nA else np.nan)
    viol_s.append(sum(1 for i in idx if pairs[i][2]['action'] == 'S'
                      and dec[i] != 'suppress'))
def ci(a):
    a = [x for x in a if not np.isnan(x)]
    return [round(float(np.percentile(a, 2.5)), 3),
            round(float(np.percentile(a, 97.5)), 3)]

summary = {
    'mergedGold': {'total': len(allg), 'byAction': {
        k: sum(1 for r in allg if r['action'] == k) for k in ('A', 'P', 'S')},
        'b3New': len(b3gold),
        'b3OverridesOfPrior': sum(1 for g in b3gold if g['overridesPrior']),
        'relayDeferred': [r['id'] for r in allg if r['relay']]},
    'intentHead86': {
        'auc_oof': round(float(roc_auc_score(yt, oof)), 3),
        'auc_calibrated_oof': round(float(roc_auc_score(yt, pcal)), 3),
        'brier_raw': round(float(brier_score_loss(yt, oof)), 3),
        'brier_calibrated': round(float(brier_score_loss(yt, pcal)), 3)},
    'chosen': chosen,
    'grid86': grid,
    'bootstrapCI_at_chosen': {
        'precision95CI': ci(prec_s),
        'recall95CI': ci(rec_s),
        'sViolations_median': float(np.median(viol_s))},
    'lrCoefficients': {f: round(float(c), 3)
                       for f, c in zip(FEAT, coef_model.coef_[0])},
    'userDesignInputs': [
        '多目标题默认 prefetch，是否升 activate 交由用户阈值设置（P3 完整性门的政策化）',
        'P4 repetition 应复用 JS 端 episodic→procedural 的多次提及判断门信号',
        '思维链上下文标记（let me think about / 让我想想）作为意图升级特征',
        '生活话题重复可按用户主动性偏好做主动联想——挂接个性化议题③'],
}
json.dump(summary, open(os.path.join(HERE, 'merged-gold-summary.json'), 'w',
                        encoding='utf-8'), ensure_ascii=False, indent=1)
print(json.dumps({'merged': summary['mergedGold'],
                  'intentHead': summary['intentHead86'],
                  'chosen': chosen, 'bootCI': summary['bootstrapCI_at_chosen'],
                  'coef': summary['lrCoefficients']}, ensure_ascii=False, indent=1))
