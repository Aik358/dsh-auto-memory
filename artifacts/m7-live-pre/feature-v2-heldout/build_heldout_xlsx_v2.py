# -*- coding: utf-8 -*-
"""Rebuild the held-out review workbook as v2.

v1 (heldout-review.xlsx) covered only the 53 queue rows; the 16 supplementary
counterfactuals were never sent for human review while their synthetic labels
had already been merged into heldout-final-gold.json. Per the main-Agent rule
("unreviewed labels are not gold"), this script merges ALL 69 proposed rows,
tags each with its training/parity independence class, and emits:

  heldout-proposed-all.jsonl           unified proposed set (synthetic labels)
  heldout-independence-report-v2.json  jaccard independence for all 69
  heldout-review-v2.xlsx               human review workbook (A/P/S/H/E)

Target interpreter: python/bench/.venv (has openpyxl).
"""
import json
import os
import sys

XLSX_SKILL_DIR = os.environ.get('XLSX_SKILL_DIR',
    r'C:\Users\JH Z\.zcode\cli\plugins\cache\zcode-plugins-official'
    r'\document-skills\0.1.1\skills\xlsx')
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, 'templates')]:
    if sub not in sys.path:
        sys.path.insert(0, sub)
from base import (setup_sheet, style_header_row, style_data_row,
                  auto_fit_columns, auto_fit_row_heights)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
TRAIN_DIR = os.path.join(HERE, '..', 'label-review-cal20260824-1954')


def load_jsonl(p):
    return [json.loads(l) for l in open(p, encoding='utf-8') if l.strip()]


queue = load_jsonl(os.path.join(HERE, 'heldout-review-queue.jsonl'))
supp_p = load_jsonl(os.path.join(HERE, 'heldout-supplementary-prefetch.jsonl'))
supp_s = load_jsonl(os.path.join(HERE, 'heldout-supplementary-suppress.jsonl'))

rows = [dict(r) for r in queue]
for r in supp_p:
    rows.append({**r, 'category': 'supplementary-prefetch',
                 'labelSource': 'strong-agent', 'synthetic': True})
for r in supp_s:
    rows.append({**r, 'category': 'supplementary-suppress',
                 'labelSource': 'strong-agent', 'synthetic': True})

# ---- independence classification (same rule as _heldout_audit.py) ----


def norm(t):
    return ''.join(c for c in str(t).lower()
                   if c.isalnum() or '\u4e00' <= c <= '\u9fff')


def bigrams(t):
    n = norm(t)
    return set(n[i:i + 2] for i in range(len(n) - 1)) or {n}


def jaccard(a, b):
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


train_bi = []
for fn in ('gold-confirmed.jsonl', 'gold-confirmed-cf.jsonl',
           'gold-confirmed-b3.jsonl'):
    for g in load_jsonl(os.path.join(TRAIN_DIR, fn)):
        if g.get('isGold'):
            train_bi.append((g['sampleId'], bigrams(g['queryText'])))
par_bi = [(f.get('sampleId', f.get('id', '')),
           bigrams(f.get('normalizedText', f.get('queryText', ''))))
          for f in load_jsonl(os.path.join(
              TRAIN_DIR, 'golden-parity-fixtures-v1.jsonl'))]

JAC = 0.5
for r in rows:
    sb = bigrams(r['queryText'])
    bt = max((jaccard(sb, tb) for _, tb in train_bi), default=0.0)
    bp = max((jaccard(sb, pb) for _, pb in par_bi), default=0.0)
    if bp >= JAC:
        r['independence'] = 'overlaps-parity'
    elif bt >= JAC:
        r['independence'] = 'overlaps-training'
    else:
        r['independence'] = 'independent-heldout'
    r['bestTrainJ'] = round(bt, 3)
    r['bestParJ'] = round(bp, 3)

with open(os.path.join(HERE, 'heldout-proposed-all.jsonl'), 'w',
          encoding='utf-8') as f:
    for r in rows:
        f.write(json.dumps(r, ensure_ascii=False) + '\n')
json.dump([{'sampleId': r['sampleId'], 'classification': r['independence'],
            'bestTrainJ': r['bestTrainJ'], 'bestParJ': r['bestParJ']}
           for r in rows],
          open(os.path.join(HERE, 'heldout-independence-report-v2.json'),
               'w'), ensure_ascii=False, indent=1)

# ---- workbook ----
ORDER = ['echo-vs-recall', 'failure-vs-planning', 'low-info', 'cross-lingual',
         'cross-workspace', 'supersede', 'supplementary-prefetch',
         'supplementary-suppress']
CAT_ZH = {'echo-vs-recall': '生活echo vs 回忆',
          'failure-vs-planning': '失败汇报 vs 计划',
          'low-info': '低信息量', 'cross-lingual': '跨语言',
          'cross-workspace': '跨工作区', 'supersede': '已被取代',
          'supplementary-prefetch': '补充·预取',
          'supplementary-suppress': '补充·抑制'}
INDEP_ZH = {'independent-heldout': '独立',
            'overlaps-parity': '与parity相似*',
            'overlaps-training': '与训练集相似*'}


def short_id(i):
    return i if len(i) <= 26 else i[:24] + '…'


rows.sort(key=lambda r: (ORDER.index(r.get('category', '?')), r['sampleId']))

wb = Workbook()
ws = wb.active
ws.title = 'Held-out 审批表'
headers = ['序号', '样本ID', '类别', '语言', '独立性', '查询文本',
           '建议\n(合成标签)', '预期目标', '生成理由', '用户选择',
           '备注\n(E时填目标ID)']
last_col = len(headers) + 1
setup_sheet(ws, title='M7 Activation v2 Held-out 独立审批表（69 条 · '
                      '含 16 条补充集）', last_col=last_col)
c3 = ws.cell(row=3, column=2,
             value='「建议」列为强Agent合成标签，仅供参考——请独立判断。'
                   '「独立性」带 * 的条目与训练集/parity 文本高相似（仍需审核，'
                   '评估时会分层报告）。选择：A=激活 P=预取 S=抑制 '
                   'H=有害 E=修改目标（备注列填正确 memoryId）。')
c3.font = Font(name='Calibri', size=9, italic=True, color='666666')
c3.alignment = Alignment(wrap_text=True, vertical='top')
ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
ws.row_dimensions[3].height = 28

for j, h in enumerate(headers):
    ws.cell(row=4, column=2 + j, value=h)
style_header_row(ws, row_num=4, col_start=2, col_end=last_col)

for i, r in enumerate(rows):
    row = 5 + i
    exp = ', '.join(short_id(x) for x in (r.get('expectedMemoryIds') or []))
    vals = [i + 1, r['sampleId'], CAT_ZH.get(r.get('category'), r.get('category')),
            r.get('language', ''), INDEP_ZH.get(r['independence'],
                                                r['independence']),
            r['queryText'], r.get('proposedAction', ''), exp or '—',
            (r.get('rationale') or '').strip() or '—', '', '']
    for j, v in enumerate(vals):
        ws.cell(row=row, column=2 + j, value=v)
    style_data_row(ws, row_num=row, col_start=2, col_end=last_col, row_index=i)
last_data_row = 4 + len(rows)

dv = DataValidation(type='list', formula1='"A,P,S,H,E"', allow_blank=True)
dv.error = '只能填 A / P / S / H / E'
dv.errorTitle = '无效选择'
ws.add_data_validation(dv)
choice_col = 2 + headers.index('用户选择')
dv.add('%s5:%s%d' % (chr(64 + choice_col), chr(64 + choice_col),
                     last_data_row))

auto_fit_columns(ws, min_width=8, max_width=44, header_row=4,
                 data_start_row=5)
auto_fit_row_heights(ws, header_row=4, data_start_row=5)

# Sheet 2: full target-id appendix (E edits need complete ids)
ws2 = wb.create_sheet('目标ID附录')
h2 = ['样本ID', 'expectedMemoryIds（完整）', 'forbiddenMemoryIds（完整）']
setup_sheet(ws2, title='目标 ID 附录（E 编辑时复制使用）', last_col=len(h2) + 1)
for j, h in enumerate(h2):
    ws2.cell(row=4, column=2 + j, value=h)
style_header_row(ws2, row_num=4, col_start=2, col_end=len(h2) + 1)
for i, r in enumerate(rows):
    row = 5 + i
    vals = [r['sampleId'],
            '\n'.join(r.get('expectedMemoryIds') or []) or '—',
            '\n'.join(r.get('forbiddenMemoryIds') or []) or '—']
    for j, v in enumerate(vals):
        ws2.cell(row=row, column=2 + j, value=v)
    style_data_row(ws2, row_num=row, col_start=2, col_end=len(h2) + 1,
                   row_index=i)
auto_fit_columns(ws2, min_width=10, max_width=52, header_row=4,
                 data_start_row=5)
auto_fit_row_heights(ws2, header_row=4, data_start_row=5)

out = os.path.join(HERE, 'heldout-review-v2.xlsx')
wb.save(out)

from collections import Counter
print('written:', out)
print('rows:', len(rows))
print('by action:', dict(Counter(r.get('proposedAction') for r in rows)))
print('by independence:', dict(Counter(r['independence'] for r in rows)))
print('lang x action:', dict(Counter((r.get('language'),
                                      r.get('proposedAction')) for r in rows)))
